"""Tool implementations for drive_mcp_extended."""
import base64
import datetime
import io
import json
from typing import Optional

from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaInMemoryUpload

from .drive_auth import get_credentials

_GOOGLE_NATIVE_EXPORT = {
    "application/vnd.google-apps.document":     "text/plain",
    "application/vnd.google-apps.spreadsheet":  "text/csv",
    "application/vnd.google-apps.presentation": "text/plain",
    "application/vnd.google-apps.form":         "text/plain",
}
_PLAIN_TEXT_TYPES = {
    "text/plain", "text/csv", "text/markdown", "application/json",
    "application/xml", "text/html", "text/javascript",
}
_IMAGE_TYPES = {"image/png", "image/jpeg", "image/jpg", "image/gif", "image/webp"}
_XLSX_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
}


# ── read_any_file ──────────────────────────────────────────────────────────────

def read_any_file(file_id: str, mime_type_hint: Optional[str] = None) -> str:
    """Read any Drive file and return extracted text/content as JSON."""
    creds = get_credentials()
    service = build("drive", "v3", credentials=creds)

    meta = service.files().get(
        fileId=file_id, fields="id,name,mimeType,size"
    ).execute()

    mime = mime_type_hint or meta.get("mimeType", "application/octet-stream")
    result = {
        "fileId": file_id,
        "title": meta.get("name", ""),
        "mimeType": mime,
        "content": "",
        "pages": None,
        "sheets": None,
        "warnings": [],
    }

    if mime == "application/vnd.google-apps.folder":
        result["content"] = ""
        result["warnings"].append("fileId is a folder, not a file — use list_folder_contents instead")
        return json.dumps({k: v for k, v in result.items() if v is not None}, ensure_ascii=False)

    if mime in _GOOGLE_NATIVE_EXPORT:
        export_mime = _GOOGLE_NATIVE_EXPORT[mime]
        data = service.files().export(fileId=file_id, mimeType=export_mime).execute()
        result["content"] = data.decode("utf-8") if isinstance(data, bytes) else data

    elif mime == "application/pdf":
        raw = _download_bytes(service, file_id)
        text, pages, warnings = _extract_pdf(raw)
        result["content"] = text
        result["pages"] = pages
        result["warnings"].extend(warnings)

    elif mime in _XLSX_TYPES:
        raw = _download_bytes(service, file_id)
        sheets, content, warnings = _extract_xlsx(raw)
        result["content"] = content
        result["sheets"] = sheets
        result["warnings"].extend(warnings)

    elif mime in _IMAGE_TYPES:
        raw = _download_bytes(service, file_id)
        result["content"] = base64.b64encode(raw).decode("ascii")

    elif mime in _PLAIN_TEXT_TYPES or mime.startswith("text/"):
        raw = _download_bytes(service, file_id)
        result["content"] = raw.decode("utf-8", errors="replace")

    else:
        raw = _download_bytes(service, file_id)
        try:
            result["content"] = raw.decode("utf-8")
        except UnicodeDecodeError:
            result["content"] = raw[:500].hex()
            result["warnings"].append(
                "Binary file — UTF-8 decode failed. First 500 bytes as hex."
            )

    return json.dumps({k: v for k, v in result.items() if v is not None}, ensure_ascii=False)


# ── list_folder_contents ───────────────────────────────────────────────────────

def list_folder_contents(
    folder_id: str,
    recursive: bool = False,
    file_types_filter: Optional[list] = None,
    _depth: int = 0,
) -> str:
    """List files in a Drive folder. Returns JSON array. Recursive max depth 3."""
    creds = get_credentials()
    service = build("drive", "v3", credentials=creds)
    files = _list_folder(service, folder_id, recursive, file_types_filter, _depth)
    return json.dumps({"folderId": folder_id, "files": files}, ensure_ascii=False)


def _list_folder(service, folder_id: str, recursive: bool, file_types_filter, depth: int) -> list:
    if depth > 3:
        return []

    _NATIVE_EXT_MAP = {
        "gdoc":   "application/vnd.google-apps.document",
        "gsheet": "application/vnd.google-apps.spreadsheet",
        "gslide": "application/vnd.google-apps.presentation",
    }
    cutoff = datetime.datetime.now(datetime.timezone.utc) - datetime.timedelta(days=7)
    results = []
    page_token = None

    while True:
        resp = service.files().list(
            q=f"'{folder_id}' in parents and trashed=false",
            fields="nextPageToken, files(id,name,mimeType,size,modifiedTime,parents)",
            pageSize=100,
            pageToken=page_token,
        ).execute()

        for f in resp.get("files", []):
            is_folder = f["mimeType"] == "application/vnd.google-apps.folder"
            if is_folder:
                if recursive:
                    results.extend(_list_folder(service, f["id"], recursive, file_types_filter, depth + 1))
                continue

            if file_types_filter:
                ext = f["name"].rsplit(".", 1)[-1].lower() if "." in f["name"] else ""
                matched = (
                    ext in file_types_filter
                    or f["mimeType"] in [_NATIVE_EXT_MAP.get(t, "") for t in file_types_filter]
                )
                if not matched:
                    continue

            recent = False
            mod_time = f.get("modifiedTime", "")
            if mod_time:
                try:
                    dt = datetime.datetime.fromisoformat(mod_time.replace("Z", "+00:00"))
                    recent = dt > cutoff
                except ValueError:
                    pass

            results.append({
                "fileId": f["id"],
                "title": f["name"],
                "mimeType": f["mimeType"],
                "size": f.get("size"),
                "modifiedTime": mod_time,
                "parents": f.get("parents", []),
                "recent": recent,
            })

        page_token = resp.get("nextPageToken")
        if not page_token:
            break

    return results


# ── update_file ────────────────────────────────────────────────────────────────

def update_file(
    file_id: str,
    text_content: Optional[str] = None,
    base64_content: Optional[str] = None,
    content_mime_type: str = "text/plain",
    title: Optional[str] = None,
) -> str:
    """Update an existing Drive file in place. Never creates a new file."""
    creds = get_credentials()
    service = build("drive", "v3", credentials=creds)

    try:
        existing = service.files().get(
            fileId=file_id, fields="id,name,mimeType,modifiedTime"
        ).execute()
    except Exception as e:
        return json.dumps({"success": False, "error": f"File not found: {e}"})

    pre_modified = existing.get("modifiedTime")
    file_meta = {}
    if title:
        file_meta["name"] = title

    if text_content is None and base64_content is None:
        if not file_meta:
            return json.dumps({"success": False, "error": "No content or title provided"})
        updated = service.files().update(
            fileId=file_id, body=file_meta, fields="id,name,modifiedTime"
        ).execute()
        return json.dumps({
            "success": True,
            "fileId": updated["id"],
            "title": updated.get("name"),
            "modifiedTime": updated.get("modifiedTime"),
            "writeConfirmed": updated.get("modifiedTime") != pre_modified,
        })

    raw = (
        text_content.encode("utf-8")
        if text_content is not None
        else base64.b64decode(base64_content)
    )

    media = MediaInMemoryUpload(raw, mimetype=content_mime_type, resumable=False)
    try:
        updated = service.files().update(
            fileId=file_id,
            body=file_meta,
            media_body=media,
            fields="id,name,mimeType,modifiedTime",
        ).execute()
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})

    return json.dumps({
        "success": True,
        "fileId": updated["id"],
        "title": updated.get("name"),
        "modifiedTime": updated.get("modifiedTime"),
        "writeConfirmed": updated.get("modifiedTime") != pre_modified,
    })


# ── Helpers ────────────────────────────────────────────────────────────────────

def _download_bytes(service, file_id: str) -> bytes:
    buf = io.BytesIO()
    req = service.files().get_media(fileId=file_id)
    dl = MediaIoBaseDownload(buf, req)
    done = False
    while not done:
        _, done = dl.next_chunk()
    return buf.getvalue()


def _extract_pdf(raw: bytes) -> tuple:
    warnings = []
    try:
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(raw))
        pages = len(reader.pages)
        parts = []
        for i, page in enumerate(reader.pages):
            text = page.extract_text() or ""
            if not text.strip():
                warnings.append(f"Page {i + 1}: no extractable text (likely image-only)")
            parts.append(text)
        return "\n\n".join(parts), pages, warnings
    except Exception as e:
        return "", 0, [f"PDF extraction failed: {e}"]


def _extract_xlsx(raw: bytes) -> tuple:
    warnings = []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        sheets = wb.sheetnames
        parts = []
        for name in sheets:
            ws = wb[name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append("\t".join(str(c) if c is not None else "" for c in row))
            if not any(r.strip() for r in rows):
                warnings.append(f"Sheet '{name}': no content")
            parts.append(f"=== {name} ===\n" + "\n".join(rows))
        return sheets, "\n\n".join(parts), warnings
    except Exception as e:
        return [], "", [f"XLSX extraction failed: {e}"]
