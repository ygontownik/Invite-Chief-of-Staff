"""
create_deal_template.py — Generate TCIP_Deal_Model_Template.xlsx
================================================================
Run once to create the 5-tab model template every new deal starts from.

    python3 create_deal_template.py

Produces: TCIP_Deal_Model_Template.xlsx (same folder as this script)

Tab structure per Deal Modeling & Presentation Standards:
  Inputs   (blue header)  — every assumption, blue font, notes in col C
  Engine   (green header) — all calculations, references Inputs only
  Outputs  (gold header)  — one section per deck slide, all live formulas
  Waterfall (grey header) — supporting waterfall detail
  Checks   (light green)  — integrity checks, all must be green before deck goes out
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

OUT = Path(__file__).parent / "TCIP_Deal_Model_Template.xlsx"

# ── Colors ────────────────────────────────────────────────────────────────────
TAB_BLUE       = "1F4E79"   # Inputs tab
TAB_GREEN      = "375623"   # Engine tab
TAB_GOLD       = "7F6000"   # Outputs tab
TAB_GREY       = "595959"   # Waterfall tab
TAB_LTGREEN    = "4EA72A"   # Checks tab

HDR_BLUE_FILL  = PatternFill("solid", fgColor="BDD7EE")
HDR_GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
HDR_GOLD_FILL  = PatternFill("solid", fgColor="FFE699")
HDR_GREY_FILL  = PatternFill("solid", fgColor="D9D9D9")
HDR_LTGRN_FILL = PatternFill("solid", fgColor="E2EFDA")
ROW_ALT_FILL   = PatternFill("solid", fgColor="F2F2F2")
GREEN_FILL     = PatternFill("solid", fgColor="C6EFCE")
RED_FILL       = PatternFill("solid", fgColor="FFC7CE")

BLUE_FONT    = Font(color="1F4E79", bold=False)
BOLD_BLUE    = Font(color="1F4E79", bold=True)
BLACK_FONT   = Font(color="000000")          # formulas
GREEN_FONT   = Font(color="375623")          # cross-sheet links
RED_FONT     = Font(color="FF0000")          # external links
WHITE_BOLD   = Font(color="FFFFFF", bold=True)
GREY_ITALIC  = Font(color="808080", italic=True)

thin = Side(style="thin", color="CCCCCC")
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def hdr_row(ws, row, text, fill, col_span=6):
    ws.merge_cells(f"A{row}:{get_column_letter(col_span)}{row}")
    c = ws.cell(row, 1, text)
    c.font = WHITE_BOLD
    c.fill = fill
    c.alignment = CENTER
    ws.row_dimensions[row].height = 18


def col_hdr(ws, row, labels, fill):
    for ci, lbl in enumerate(labels, 1):
        c = ws.cell(row, ci, lbl)
        c.font = WHITE_BOLD
        c.fill = fill
        c.alignment = CENTER
        c.border = THIN_BORDER


def note_cell(ws, row, col, text):
    c = ws.cell(row, col, text)
    c.font = GREY_ITALIC
    c.alignment = LEFT


def input_row(ws, row, name, label, default, unit="", note="", named_range=None, wb=None):
    ws.cell(row, 1, label).font = BLACK_FONT
    c = ws.cell(row, 2, default)
    c.font = BLUE_FONT
    c.border = THIN_BORDER
    if unit:
        ws.cell(row, 3, unit).font = GREY_ITALIC
    if note:
        note_cell(ws, row, 4, note)
    if named_range and wb:
        ws_title = ws.title.replace("'", "''")
        col_letter = get_column_letter(2)
        defn = DefinedName(named_range, attr_text=f"'{ws_title}'!${col_letter}${row}")
        wb.defined_names[named_range] = defn


def check_row(ws, row, label, formula, pass_text="PASS", fail_text="FAIL"):
    ws.cell(row, 1, label).font = BLACK_FONT
    ws.cell(row, 1).alignment = LEFT
    c = ws.cell(row, 2, f"=IF({formula},\"{pass_text}\",\"{fail_text}\")")
    c.font = BLACK_FONT
    c.border = THIN_BORDER
    c.alignment = CENTER
    # Conditional formatting handled manually; leave green as placeholder
    note_cell(ws, row, 3, formula)


# ── Build workbook ────────────────────────────────────────────────────────────

def build():
    wb = Workbook()

    # ── INPUTS ────────────────────────────────────────────────────────────────
    ws_in = wb.active
    ws_in.title = "Inputs"
    ws_in.sheet_properties.tabColor = TAB_BLUE
    ws_in.column_dimensions["A"].width = 38
    ws_in.column_dimensions["B"].width = 14
    ws_in.column_dimensions["C"].width = 12
    ws_in.column_dimensions["D"].width = 40

    hdr_row(ws_in, 1, "INPUTS — every assumption lives here. Blue font = user input. Notes in col D.", HDR_BLUE_FILL, 4)
    hdr_row(ws_in, 2, "Color convention: Blue = input  |  Black = formula  |  Green = cross-sheet link  |  Red = external link", HDR_BLUE_FILL, 4)

    ws_in.cell(3, 1, "Assumption").font = BOLD_BLUE
    ws_in.cell(3, 2, "Value").font     = BOLD_BLUE
    ws_in.cell(3, 3, "Unit").font      = BOLD_BLUE
    ws_in.cell(3, 4, "Note").font      = BOLD_BLUE

    sections = [
        ("DEAL PARAMETERS", [
            ("HOLD_YEARS",      "Hold period",          5,      "years",  "Investment horizon for IRR calc"),
            ("GROSS_IRR",       "Gross IRR assumption", 0.15,   "decimal","15% = 0.15"),
            ("HURDLE_RATE",     "Hurdle rate",          0.08,   "decimal","Profit-only pref. 8% = 0.08"),
        ]),
        ("FUND ECONOMICS", [
            ("FUND_SIZE",       "Fund target size",     500,    "$M",     "Committed capital"),
            ("MGMT_FEE_PCT",    "Management fee %",     0.015,  "decimal","1.5% = 0.015"),
            ("CARRY_PCT",       "Carried interest %",   0.20,   "decimal","20% = 0.20"),
            ("LP_PCT",          "LP co-invest %",       0.10,   "decimal","Optional LP co-invest"),
        ]),
        ("FRE / MARGIN", [
            ("FRE_MULTIPLE",    "FRE terminal multiple",15,     "x",      "Applied to Yr5 run-rate FRE"),
            ("FRE_MARGIN_500",  "FRE margin at $500M",  0.30,   "decimal","30% = 0.30"),
            ("FRE_MARGIN_1B",   "FRE margin at $1B",    0.40,   "decimal","40% = 0.40"),
            ("FRE_MARGIN_2B",   "FRE margin at $2B+",   0.50,   "decimal","50% = 0.50"),
        ]),
        ("GP FORMULA (delete if not applicable)", [
            ("GP_ANCHOR",       "GP% at anchor",        0.50,   "decimal","Formula output at base case"),
            ("GP_ALPHA",        "Asset exponent (α)",   0.40,   "",       "Sensitivity of GP% to asset AUM"),
            ("GP_BETA",         "Raise exponent (β)",   0.35,   "",       "Sensitivity of GP% to raise"),
            ("GP_FLOOR",        "GP% floor",            0.25,   "decimal","Minimum GP% regardless of formula"),
            ("GP_CAP",          "GP% cap",              0.65,   "decimal","Maximum GP% regardless of formula"),
        ]),
        ("G&A / COSTS", [
            ("GNA_GROSS",       "Annual firm G&A",      2.5,    "$M/yr",  "Gross operating costs"),
            ("CEO_PAYMENT",     "CEO/ops payment",      1.0,    "$M/yr",  "Fixed management payment"),
        ]),
    ]

    row = 4
    for section_title, items in sections:
        ws_in.cell(row, 1, section_title).font = Font(bold=True, color="1F4E79")
        ws_in.cell(row, 1).fill = HDR_BLUE_FILL
        ws_in.merge_cells(f"A{row}:D{row}")
        row += 1
        for named_range, label, default, unit, note in items:
            input_row(ws_in, row, named_range, label, default, unit, note, named_range, wb)
            row += 1
        row += 1

    ws_in.freeze_panes = "A4"

    # ── ENGINE ────────────────────────────────────────────────────────────────
    ws_eng = wb.create_sheet("Engine")
    ws_eng.sheet_properties.tabColor = TAB_GREEN
    ws_eng.column_dimensions["A"].width = 40
    ws_eng.column_dimensions["B"].width = 18
    ws_eng.column_dimensions["C"].width = 40

    hdr_row(ws_eng, 1, "ENGINE — all calculations. References Inputs only. No hardcoded values.", HDR_GREEN_FILL, 3)

    rows_eng = [
        ("WATERFALL", None, None),
        ("Hold years",        "=Inputs!HOLD_YEARS",    ""),
        ("Gross IRR",         "=Inputs!GROSS_IRR",     ""),
        ("Hurdle rate",       "=Inputs!HURDLE_RATE",   ""),
        ("", None, None),
        ("Gross exit value",  "=Inputs!FUND_SIZE*(1+Inputs!GROSS_IRR)^Inputs!HOLD_YEARS", "$M"),
        ("Fee drag",          "=Inputs!FUND_SIZE*Inputs!MGMT_FEE_PCT*Inputs!HOLD_YEARS",  "$M  (fees paid over hold)"),
        ("Net proceeds",      "=B7-B8",                "$M  (exit - fee drag - capital return - use full formula)"),
        ("LP pref (profit-only)", "=Inputs!FUND_SIZE*((1+Inputs!HURDLE_RATE)^Inputs!HOLD_YEARS-1)", "$M  PROFIT ONLY — not full stack"),
        ("Catchup cap",       "=Inputs!CARRY_PCT/(1-Inputs!CARRY_PCT)*B10", "$M"),
        ("Above-hurdle pool", "=MAX(0,B8-B10)",        "$M  net proceeds - LP pref"),
        ("Carry (full catchup)","=Inputs!CARRY_PCT*MAX(0,B8-Inputs!FUND_SIZE*((1+Inputs!HURDLE_RATE)^Inputs!HOLD_YEARS-1))", "$M  = carry% × net profit when full catchup"),
        ("", None, None),
        ("FRE SCHEDULE", None, None),
        ("FRE margin (current)", "=IF(Inputs!FUND_SIZE<=500,Inputs!FRE_MARGIN_500,IF(Inputs!FUND_SIZE<=1000,Inputs!FRE_MARGIN_500+(Inputs!FRE_MARGIN_1B-Inputs!FRE_MARGIN_500)*(Inputs!FUND_SIZE-500)/500,IF(Inputs!FUND_SIZE<=2000,Inputs!FRE_MARGIN_1B+(Inputs!FRE_MARGIN_2B-Inputs!FRE_MARGIN_1B)*(Inputs!FUND_SIZE-1000)/1000,Inputs!FRE_MARGIN_2B)))", "tiered by fund size"),
        ("Yr5 FRE",           "=Inputs!FUND_SIZE*Inputs!MGMT_FEE_PCT*B17", "$M  = fund fee × margin"),
        ("Terminal value",    "=B18*Inputs!FRE_MULTIPLE",  "$M  = Yr5 FRE × multiple"),
    ]

    r = 2
    for item in rows_eng:
        lbl, formula, note = item
        if formula is None:
            if lbl:
                ws_eng.cell(r, 1, lbl).font = Font(bold=True, color="375623")
                ws_eng.cell(r, 1).fill = HDR_GREEN_FILL
            r += 1
            continue
        ws_eng.cell(r, 1, lbl).font = BLACK_FONT
        c = ws_eng.cell(r, 2, formula)
        c.font = GREEN_FONT if formula.startswith("=Inputs!") else BLACK_FONT
        c.border = THIN_BORDER
        if note:
            note_cell(ws_eng, r, 3, note)
        r += 1

    # Named ranges for key Engine outputs
    for name, cell in [("CARRY",       "B13"), ("YR5_FRE", "B18"),
                       ("TERMINAL_VALUE","B19"), ("FRE_MARGIN_CURR","B17"),
                       ("NET_PROCEEDS", "B9")]:
        defn = DefinedName(name, attr_text=f"'Engine'!${cell[0]}${cell[1:]}")
        wb.defined_names[name] = defn

    # ── OUTPUTS ───────────────────────────────────────────────────────────────
    ws_out = wb.create_sheet("Outputs")
    ws_out.sheet_properties.tabColor = TAB_GOLD
    ws_out.column_dimensions["A"].width = 22

    hdr_row(ws_out, 1, "OUTPUTS — one section per deck slide. All live formulas. build_deck.py reads from here via named ranges.", HDR_GOLD_FILL, 8)

    slide_sections = [
        "SLIDE 1 — Cover / Summary",
        "SLIDE 2 — Question / Setup",
        "SLIDE 3 — Formula / Mechanism",
        "SLIDE 4 — GP% Matrix  (4 rows × 6 cols: FIT AUM × TCIP raise)",
        "SLIDE 5 — Baazov $ Matrix  (same dimensions)",
        "SLIDE 6 — TCIP $ Matrix   (same dimensions)",
        "SLIDE 7 — Waterfall Decomposition",
        "SLIDE 8 — Sensitivities",
        "SLIDE 9 — Full Economics",
        "SLIDE 10 — Term Sheet",
        "SLIDE 11 — Takeaways",
    ]

    r = 2
    for section in slide_sections:
        ws_out.cell(r, 1, section).font = Font(bold=True, color="7F6000")
        ws_out.cell(r, 1).fill = HDR_GOLD_FILL
        note_cell(ws_out, r + 1, 1, "← Insert output formulas here. Named range every cell build_deck.py reads.")
        note_cell(ws_out, r + 2, 1, "← Col headers row (labels only, no formulas needed)")
        r += 5

    # ── WATERFALL ─────────────────────────────────────────────────────────────
    ws_wf = wb.create_sheet("Waterfall")
    ws_wf.sheet_properties.tabColor = TAB_GREY
    ws_wf.column_dimensions["A"].width = 40
    ws_wf.column_dimensions["B"].width = 16
    ws_wf.column_dimensions["C"].width = 16
    ws_wf.column_dimensions["D"].width = 40

    hdr_row(ws_wf, 1, "WATERFALL — supporting detail. Not read directly by build_deck.py.", HDR_GREY_FILL, 4)

    wf_rows = [
        ("LP PREF BASE", None, None, "PROFIT-ONLY (standard PE). Ask explicitly before building."),
        ("Capital invested",     "=Inputs!FUND_SIZE", "$M", ""),
        ("Gross exit",           "=Inputs!FUND_SIZE*(1+Inputs!GROSS_IRR)^Inputs!HOLD_YEARS", "$M", ""),
        ("Gross profit",         "=B3-B2", "$M", ""),
        ("LP pref (profit-only)","=Inputs!FUND_SIZE*((1+Inputs!HURDLE_RATE)^Inputs!HOLD_YEARS-1)", "$M", "=capital × ((1+hurdle)^hold − 1)  ← PROFIT ONLY"),
        ("Above-hurdle pool",    "=MAX(0,B4-B5)", "$M", "Gross profit − LP pref"),
        ("", None, None, ""),
        ("CARRY SPLIT (full catchup)", None, None, "Full catchup identity: carry = carry% × net profit"),
        ("Net profit (fee-adjusted)", "=B4-Engine!B8", "$M", "Gross profit − fee drag"),
        ("LP pref catchup cap",  "=Inputs!CARRY_PCT/(1-Inputs!CARRY_PCT)*B5", "$M", ""),
        ("GP catchup",           "=MIN(B10,MAX(0,B6))", "$M", "100% to GP until GP holds carry% of profits"),
        ("Above-catchup",        "=MAX(0,B6-B10)", "$M", ""),
        ("GP carry on above-catchup", "=Inputs!CARRY_PCT*B12", "$M", ""),
        ("TOTAL CARRY",          "=B11+B13", "$M", "Must equal: carry% × net profit"),
        ("", None, None, ""),
        ("CROSS-CHECK",          "=B14", "$M", "Should equal: =Inputs!CARRY_PCT*B9"),
        ("Difference (must be 0)","=B14-Inputs!CARRY_PCT*B9","",  "If non-zero, waterfall formula is wrong"),
    ]

    r = 2
    for row_data in wf_rows:
        lbl, val, unit, note = row_data
        if val is None:
            if lbl:
                ws_wf.cell(r, 1, lbl).font = Font(bold=True, color="595959")
                ws_wf.cell(r, 1).fill = HDR_GREY_FILL
            r += 1
            continue
        ws_wf.cell(r, 1, lbl).font = BLACK_FONT
        if val:
            c = ws_wf.cell(r, 2, val)
            c.font = GREEN_FONT if "Inputs!" in val else BLACK_FONT
            c.border = THIN_BORDER
            c.number_format = '#,##0.0'
        if unit:
            ws_wf.cell(r, 3, unit).font = GREY_ITALIC
        if note:
            note_cell(ws_wf, r, 4, note)
        r += 1

    # ── CHECKS ────────────────────────────────────────────────────────────────
    ws_chk = wb.create_sheet("Checks")
    ws_chk.sheet_properties.tabColor = TAB_LTGREEN
    ws_chk.column_dimensions["A"].width = 45
    ws_chk.column_dimensions["B"].width = 10
    ws_chk.column_dimensions["C"].width = 55

    hdr_row(ws_chk, 1, "CHECKS — all must show PASS before deck goes out. Fix failures before running build_deck.py.", HDR_LTGRN_FILL, 3)
    col_hdr(ws_chk, 2, ["Check", "Result", "Formula / How to verify"], HDR_LTGRN_FILL)

    checks = [
        ("Carry = carry% × net profit",
         "ABS(Engine!B13-Inputs!CARRY_PCT*Waterfall!B9)<0.5",
         "Mathematical identity when full catchup + profit-only pref. If fails, waterfall formula is wrong."),
        ("GP stake = FRE + carry + terminal",
         "ABS(GP_STAKE_TOTAL-(CUMUL_FRE+CARRY+TERMINAL_VALUE))<0.5",
         "Components must sum to total. Add named ranges for your deal's components."),
        ("TCIP total = GP stake share + promote",
         "ABS(TCIP_TOTAL-(TCIP_GP_STAKE+PROMOTE))<0.5",
         "Add named ranges for your deal's TCIP total and components."),
        ("Yr5 FRE = fund fee × FRE margin",
         "ABS(YR5_FRE-Inputs!FUND_SIZE*Inputs!MGMT_FEE_PCT*FRE_MARGIN_CURR)<0.1",
         "FRE cross-check. Confirms margin tier logic is applied correctly."),
        ("No #REF! or #VALUE! errors in Outputs",
         "ISERROR(Outputs!B5)=FALSE",
         "Spot-check. Verify manually that Outputs tab has no error cells before build."),
        ("All named ranges resolve (manual)",
         "TRUE",
         "Run: python3 -c \"from build_deck_<deal> import read_fit_model; print('OK')\""),
    ]

    r = 3
    for label, formula, note in checks:
        ws_chk.cell(r, 1, label).font = BLACK_FONT
        ws_chk.cell(r, 1).alignment = LEFT
        c = ws_chk.cell(r, 2, f"=IF({formula},\"PASS\",\"FAIL\")")
        c.font = BLACK_FONT
        c.border = THIN_BORDER
        c.alignment = CENTER
        note_cell(ws_chk, r, 3, note)
        r += 1

    r += 1
    ws_chk.cell(r, 1, "Note: add conditional formatting to column B — green fill for PASS, red fill for FAIL.").font = GREY_ITALIC

    # ── Tab order ─────────────────────────────────────────────────────────────
    # openpyxl sheet order is creation order; already correct

    wb.save(str(OUT))
    print(f"✓ Template written: {OUT}")
    print(f"  Tabs: {[s.title for s in wb.worksheets]}")
    print(f"  Named ranges: {len(list(wb.defined_names))}")
    print()
    print("Next steps:")
    print("  1. Copy this file → rename to <deal>_Model_v1.xlsx")
    print("  2. Populate Inputs tab (blue cells). Build Engine + Outputs.")
    print("  3. Standards doc: https://docs.google.com/document/d/16VVM5k_nvZMAXKr-sEvnTpnoASgabq7z82BeO0qO0IA/edit")


if __name__ == "__main__":
    build()
